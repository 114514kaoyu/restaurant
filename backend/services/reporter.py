"""
报告导出服务 (基础版: Excel)
PDF 导出将在第三批实现
"""
import io
import logging
from sqlalchemy.orm import Session

from services.analyzer import DataAnalyzer
from services.ranker import RankingService

logger = logging.getLogger(__name__)


class ReportService:

    def __init__(self, db: Session):
        self.db = db
        self.analyzer = DataAnalyzer(db)
        self.ranker = RankingService(db)

    def export_excel(self, snapshot_id: int, algorithm: str = "wilson") -> io.BytesIO:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()

        # ── Sheet 1: 排名总表 ──
        ws1 = wb.active
        ws1.title = "餐厅排名"
        headers = [
            "排名", "名称", "菜系", "评分", "评价数",
            "人均消费", "区域", "地址", "算法得分",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        data = self.ranker.get_ranking_list(
            snapshot_id=snapshot_id,
            algorithm=algorithm,
            page=1,
            page_size=10000,
        )
        for row_idx, item in enumerate(data["items"], 2):
            ws1.cell(row=row_idx, column=1, value=item["rank"])
            ws1.cell(row=row_idx, column=2, value=item["name"])
            ws1.cell(row=row_idx, column=3, value=item.get("category", ""))
            ws1.cell(row=row_idx, column=4, value=item.get("rating"))
            ws1.cell(row=row_idx, column=5, value=item.get("rating_count"))
            ws1.cell(row=row_idx, column=6, value=item.get("cost_avg"))
            ws1.cell(row=row_idx, column=7, value=item.get("district", ""))
            ws1.cell(row=row_idx, column=8, value=item.get("address", ""))
            ws1.cell(row=row_idx, column=9, value=round(item["score"], 4))

        # 自动列宽
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Sheet 2: 总览统计 ──
        ws2 = wb.create_sheet("总览统计")
        overview = self.analyzer.overview(snapshot_id)
        label_map = {
            "total": "餐厅总数",
            "avg_rating": "平均评分",
            "median_rating": "中位数评分",
            "max_rating": "最高评分",
            "min_rating": "最低评分",
            "avg_price": "平均人均消费",
            "total_categories": "菜系数量",
            "total_districts": "区域数量",
            "avg_reviews": "平均评价数",
        }
        for i, (key, value) in enumerate(overview.items(), 1):
            ws2.cell(row=i, column=1, value=label_map.get(key, key))
            ws2.cell(row=i, column=2, value=value)

        # ── Sheet 3: 菜系统计 ──
        ws3 = wb.create_sheet("菜系统计")
        cat_headers = ["菜系", "数量", "平均评分", "平均价格", "最高评分", "总评价数"]
        for col, h in enumerate(cat_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for i, item in enumerate(self.analyzer.category_stats(snapshot_id), 2):
            ws3.cell(row=i, column=1, value=item["category"])
            ws3.cell(row=i, column=2, value=item["count"])
            ws3.cell(row=i, column=3, value=item["avg_rating"])
            ws3.cell(row=i, column=4, value=item["avg_price"])
            ws3.cell(row=i, column=5, value=item["max_rating"])
            ws3.cell(row=i, column=6, value=item["total_reviews"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output